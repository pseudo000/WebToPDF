# 헤들리스 브라우저에서 png->pdf변환까지는 완료
# 프로그레스바 활성화 
# 실패메시지 출력위젯 추가
# 작업 메모리 줄이기
# merged시 공백제거 

import tkinter.ttk as ttk   
from tkinter import *  #__all__? filedialog는 서브 모듈이기 때문에 별도 임포트 해줘야 됨 
from tkinter import filedialog
import openpyxl
import os
import PyPDF2
import re
import threading
from selenium import webdriver
from PIL import Image
import io
import os
from openpyxl import load_workbook
from collections import defaultdict
from shutil import copyfile
from PyPDF2 import PdfMerger
import tkinter.messagebox as mbox
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
from selenium.common.exceptions import TimeoutException
from PyPDF2 import PdfFileReader, PdfFileWriter
from fpdf import FPDF
from collections import defaultdict
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver


root = Tk()
root.title("SP IMPORT")
root.resizable(False,False)
root.geometry("500x600")


# 엑셀파일의 a열에 (1)...을 미리 입력 후 해당 값으로 pdf 파일 보존
# pdf파일 원본이름 그대로 대조하여 excel check 테스트


#기본 다운로드 경로 선택 
def add_path():
    addpath_selected = filedialog.askdirectory()
    if addpath_selected == '': #사용자가 취소를 누를 때
        return
    txt_frame1_path.delete(0, END)
    txt_frame1_path.insert(0,addpath_selected)
    return addpath_selected


#파일추가
def add_file():
    files = filedialog.askopenfilename(title="xlsx 파일을 선택하세요", \
        filetypes=(("xlsx 파일", "*.xlsx"),("All files", "*.*")), \
        initialdir="C:/")
    #사용자가 선택한 파일출력
    for file in files:
        txt_xls_path.insert(END, file)
    return files[0]

#MergedPDF저장경로
def savemerged_path():
    savemerged_selected = filedialog.askdirectory()
    if savemerged_selected == '': #사용자가 취소를 누를 때
        return
    txt_mergedpdf_path.delete(0, END)
    txt_mergedpdf_path.insert(0,savemerged_selected)
    return savemerged_selected


def convert_png_to_pdf(png_path, pdf_path):
    # PNG 이미지를 열고 크기를 가져옵니다
    image = Image.open(png_path)

    # 이미지 크기를 원하는 크기로 조정합니다
    target_width = 210  # 210mm
    target_height = 297  # 297mm
    image = image.resize((int(target_width), int(target_height)))

    # PDF를 생성할 FPDF 객체를 생성합니다
    pdf = FPDF(unit="mm", format=(target_width, target_height))  # A4 용지 크기

    # PDF에 이미지를 추가합니다
    pdf.add_page()
    pdf.image(png_path, x=0, y=0, w=target_width, h=target_height)

    # PDF 파일을 저장합니다
    pdf.output(pdf_path, "F")



def WebToPDF(url, file_name):
    DRIVER_PATH = 'chromedriver.exe'

    options = Options()
    # options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--start-minimized')
    service = Service(DRIVER_PATH)

    wait_time = 5
    # driver = webdriver.Chrome(DRIVER_PATH, options=options)
    driver = webdriver.Chrome(service=service, options=options)  # Service 객체를 사용하여 ChromeDriver 실행
    driver.set_page_load_timeout(30)  # 페이지 로드 타임아웃을 30초로 설정

    try:
        driver.get(url)
        WebDriverWait(driver, wait_time).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
    except TimeoutException:
        # 접속 대기 시간이 30초를 넘어가면 패스
        driver.quit()
        raise TimeoutException("페이지 로드 타임아웃")

    # 스크롤할 높이와 너비를 설정하세요
    scroll_height = 1000  # 세로 스크롤할 높이
    scroll_width = 500  # 가로 스크롤할 너비

    # 현재 창의 초기 크기를 저장합니다
    initial_width = driver.execute_script("return window.innerWidth")
    initial_height = driver.execute_script("return window.innerHeight")

    # 창의 크기를 스크롤할 높이와 너비만큼 조정합니다
    driver.set_window_size(initial_width + scroll_width, initial_height + scroll_height)

    time.sleep(1)
    
    # 스크롤된 화면의 스크린샷을 찍습니다
    png = driver.get_screenshot_as_png()
    im = Image.open(io.BytesIO(png)).convert('RGB')

    # 스크린샷의 해상도를 조정합니다
    dpi = 100  # 원하는 해상도
    new_width = int(initial_width * dpi / 96)
    new_height = int(initial_height * dpi / 96)
    im = im.resize((new_width, new_height), resample=Image.LANCZOS)

    # 이미지를 파일로 저장합니다
    png_path = os.path.join(txt_frame1_path.get(), file_name + '.png')
    im.save(png_path)

    # 원래 창의 크기로 되돌립니다
    driver.set_window_size(initial_width, initial_height)

    pdf_path = os.path.join(txt_frame1_path.get(), file_name + '.pdf')
    convert_png_to_pdf(png_path, pdf_path)

    driver.quit()
    os.remove(png_path)



def start_process():
    xl_file_path = txt_xls_path.get()
    workbook = openpyxl.load_workbook(xl_file_path, data_only=True)
    worksheet = workbook.worksheets[0]

    txt_frame2 = Text(root, height=10)
    txt_frame2.pack(fill="both", expand=True)
    scroll_bar = Scrollbar(txt_frame2, orient="vertical", command=txt_frame2.yview)
    scroll_bar.pack(side="right", fill="y")
    txt_frame2.config(yscrollcommand=scroll_bar.set)

    fail_count = 0 
    for row, item in enumerate(worksheet.rows):
        if row > 0:
            url = item[1].value
            file_name = item[0].value

            try:
                WebToPDF(url,file_name)

            except Exception as e:
                fail_count += 1
                message = file_name + '실패\n'
                txt_frame2.insert(END, message)

        # 진행바 업데이트
        p_var.set((row / worksheet.max_row) * 100)
        progress_bar.update()
        root.update()  # UI 업데이트 강제 수행

    p_var.set(100)
    if fail_count == 0:
        message2 = "작업이 모두 완료되었습니다."
    else:
        message2 = f"{fail_count}/{worksheet.max_row-1}개 작업에 실패하였습니다."
    txt_frame2.insert(END, message2)
    mbox.showinfo("COMPLETE", "작업이 모두 완료되었습니다.")
    p_var.set(0)
    
    
def run_thread():
    # 스레드 생성 및 시작
    thread = threading.Thread(target=start_process)
    thread.daemon = True
    thread.start()




def split_process():
# PDF 파일이 있는 폴더의 경로를 정의합니다.
    folder_path = txt_frame1_path.get()

# 폴더 내의 모든 파일을 반복합니다.
    for filename in os.listdir(folder_path):
        # PDF 파일만 처리합니다.
        if filename.endswith(".pdf"):
            # 파일의 전체 경로를 생성합니다.
            file_path = os.path.join(folder_path, filename)
            
            # PyPDF2를 사용하여 PDF 파일을 읽습니다.
            pdf_file = PyPDF2.PdfReader(file_path)
            
            # 수정된 버전을 저장할 새 PDF 파일을 생성합니다.
            output_pdf = PyPDF2.PdfWriter()
            
            # 새 PDF 파일에 1페이지를 추가합니다.
            output_pdf.add_page(pdf_file.pages[0])
        
            # 새 PDF 파일을 디스크에 쓰기
            with open(file_path, "wb") as output_file:
                output_pdf.write(output_file)


def check_pdf_excel():
    # PDF 파일이 있는 폴더 경로
    pdf_folder_path = txt_frame1_path.get()

    # 엑셀 파일 경로
    excel_file_path = txt_xls_path.get()

    # PDF 파일 이름과 경로를 딕셔너리로 저장
    pdf_dict = {}
    for pdf in os.listdir(pdf_folder_path):
        if pdf.endswith(".pdf"):
            pdf_name = os.path.splitext(pdf)[0] 
            pdf_path = os.path.join(pdf_folder_path, pdf)
            pdf_dict[pdf_name] = pdf_path

    # 엑셀 파일 열기
    workbook = load_workbook(filename=excel_file_path)
    worksheet = workbook.active

    # A 열에서 PDF 파일 이름 가져오기
    pdf_files = [cell.value for cell in worksheet['A'][1:]]

    # C 열에서 PDF 파일 이름에 해당하는 값 가져오기
    for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            pdf_name = cell.offset(column=-2).value
            if pdf_name in pdf_files:
                pdf_path = pdf_dict.get(pdf_name)
                if pdf_path:
                    cell.value = pdf_files[pdf_files.index(pdf_name)]
                else:
                    cell.value = "" 
            else:
                cell.value = ""  

    workbook.save(excel_file_path)  


def compress_pdf(input_path, output_path, compression_quality):
    pdf = PdfFileReader(input_path)
    writer = PdfFileWriter()

    for page_number in range(pdf.getNumPages()):
        page = pdf.getPage(page_number)
        page.compressContentStreams()  # 인자 제거

        writer.addPage(page)

    with open(output_path, 'wb') as output_file:
        writer.write(output_file)

def add_suffix_to_11_char_filenames(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith('.pdf'):
            file_name_without_extension = os.path.splitext(filename)[0]
            if len(file_name_without_extension) == 11:
                new_filename = file_name_without_extension + ' (0)' + '.pdf'
                old_filepath = os.path.join(folder_path, filename)
                new_filepath = os.path.join(folder_path, new_filename)
                os.rename(old_filepath, new_filepath)


def merge_process():
    # 병합할 폴더 경로와 결과물 파일이 저장될 폴더 경로 설정
    folder_path = txt_frame1_path.get()
    result_folder_path = txt_mergedpdf_path.get()

    if not os.path.exists(result_folder_path):
        os.makedirs(result_folder_path)

    add_suffix_to_11_char_filenames(folder_path)

    # 파일 이름의 앞 11자리가 동일한 파일들을 그룹화
    file_dict = defaultdict(list)
    for filename in os.listdir(folder_path):
        if filename.endswith('.pdf'):
            key = filename[:11]  # 앞 11자리까지만 추출하여 key로 사용
            file_dict[key].append(filename)

    # 그룹화된 파일들을 순회하며 각 그룹을 병합 또는 복사
    for key, filenames in file_dict.items():
        if len(filenames) == 1:
            file_path = os.path.join(folder_path, filenames[0])
            result_name = re.sub(r'\(\d+\)', '', filenames[0])  # 괄호와 괄호 안에 있는 숫자 삭제
            result_name = re.sub(r'-\d+', '', result_name)
            result_name = re.sub(r'\s+', '', result_name)  # 공백 삭제
            result_path = os.path.join(result_folder_path, result_name)
            copyfile(file_path, result_path)
            # PDF 파일 압축
            compress_pdf(result_path, result_path, compression_quality=0.5)  # 압축 품질 조정 가능 (0.1 ~ 1.0)
        else:
            merger = PyPDF2.PdfFileMerger()
            for filename in sorted(filenames):
                file_path = os.path.join(folder_path, filename)
                merger.append(file_path)
            result_name = filenames[0]
            result_name = re.sub(r'\(\d+\)', '', result_name)  # 괄호와 괄호 안에 있는 숫자 삭제
            result_name = re.sub(r'-\d+', '', result_name)
            result_name = re.sub(r'\s+', '', result_name)  # 공백 삭제
            i = 1
            while os.path.exists(os.path.join(result_folder_path, result_name)):
                result_name = re.sub(r'\(\d+\)', '', result_name) + f'({i})'
                i += 1
            result_path = os.path.join(result_folder_path, result_name)
            merger.write(result_path)
            merger.close()
            # PDF 파일 압축
            compress_pdf(result_path, result_path, compression_quality=0.5)  # 압축 품질 조정 가능 (0.1 ~ 1.0)



#프레임1 (기본 다운로드폴더 지정)
frame1 = LabelFrame(root, text="Select Download folder")
frame1.pack(fill="x", padx=5, pady=5, ipady=5)

txt_frame1_path = Entry(frame1)
txt_frame1_path.pack(side="left", fill="x", expand="True", ipady=4) #

btn_frame1_path = Button(frame1, text="...folder",padx=5, pady=5, width=12, command = add_path)
btn_frame1_path.pack(side="right")



#Web to PDF 저장
savepdf_frame = LabelFrame(root, text="Web to PDF")
savepdf_frame.pack(fill="x", padx=5, pady=5, ipady=5)

txt_xls_path = Entry(savepdf_frame)
txt_xls_path.pack(side="left", fill="x", expand="True", ipady=4) #

btn_start = Button(savepdf_frame, text="START",padx=5, pady=5, width=12, command=run_thread)
btn_start.pack(side="right")

btn_savepdf_path = Button(savepdf_frame, text="...xlsx file",padx=5, pady=5, width=12, command=add_file)
btn_savepdf_path.pack(side="right")



#Slit PDF
splitpdf_frame = LabelFrame(root, text="Split PDF")
splitpdf_frame.pack(fill="x",padx=5, pady=5, ipady=5)

btn_splitpdf = Button(splitpdf_frame, text="START", padx=5, pady=5, width="20", command=split_process)
btn_splitpdf.pack()



#Excel Check
xlscheck_frame = LabelFrame(root, text="Check PDF EXCEL")
xlscheck_frame.pack(fill="x",padx=5, pady=5, ipady=5)

btn_xlscheck = Button(xlscheck_frame, text="START", padx=5, pady=5, width="20", command=check_pdf_excel)
btn_xlscheck.pack()



# Merged PDF 저장경로
mergedpdf_frame = LabelFrame(root, text="Folder to merged PDF")
mergedpdf_frame.pack(fill="x", padx=5, pady=5, ipady=5)

txt_mergedpdf_path = Entry(mergedpdf_frame)
txt_mergedpdf_path.pack(side="left", fill="x", expand="True", ipady=4) #

btn_mergestart = Button(mergedpdf_frame, text="START",padx=5, pady=5, width=12, command=merge_process)
btn_mergestart.pack(side="right")

btn_mergedpdf_path = Button(mergedpdf_frame, text="...Folder to save",padx=5, pady=5, width=12, command=savemerged_path)
btn_mergedpdf_path.pack(side="right")





# 진행상황 profress bar
frame_progress = LabelFrame(root, text="진행상황")
frame_progress.pack(fill="x")

p_var = DoubleVar()
progress_bar = ttk.Progressbar(frame_progress, maximum=100, variable=p_var)
progress_bar.pack(fill="x", pady=10)


# 종료 프레임
close_frame = Frame(root)
close_frame.pack(side="right", padx=5, pady=5)

btn_close = Button(close_frame, padx=5, pady=5, text="CLOSE", width=12, command=root.destroy)
btn_close.pack()








root.mainloop()

# pyinstaller -w -F --add-data '*.png;.' 5_add...............
# "C:\Users\swwoo\Desktop\sp import\WebToPDF\GUIgogo\5_headless3.py"
# pyinstaller -w -F --onefile --add-binary "chromedriver.exe;." --add-binary "C:\Program Files\Google\Chrome\Application\chrome.exe;." "C:\Users\swwoo\Desktop\sp import\WebToPDF\GUIgogo\5_headless3.py"
# pyinstaller -w "C:\Users\swwoo\Desktop\sp import\WebToPDF\GUIgogo\11_sort.py"
