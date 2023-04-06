import tkinter.ttk as ttk   
from tkinter import *  #__all__? filedialog는 서브 모듈이기 때문에 별도 임포트 해줘야 됨 
from tkinter import filedialog
import json
import time
import openpyxl
import os
import shutil
import PyPDF2
import re
import threading
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from collections import defaultdict
from shutil import copyfile
from PyPDF2 import PdfMerger

root = Tk()
root.title("SP IMPORT")
root.resizable(False,False)
root.geometry("500x600")


#깃허브 병신같음
#기본 다운로드 경로 선택 
def add_path():
    addpath_selected = filedialog.askdirectory()
    if addpath_selected == '': #사용자가 취소를 누를 때
        return
    txt_frame1_path.delete(0, END)
    txt_frame1_path.insert(0,addpath_selected)
    return addpath_selected


#파일추가
def add_file():
    files = filedialog.askopenfilename(title="xlsx 파일을 선택하세요", \
        filetypes=(("xlsx 파일", "*.xlsx"),("All files", "*.*")), \
        initialdir="C:/")
    #사용자가 선택한 파일출력
    for file in files:
        txt_xls_path.insert(END, file)
    return files[0]

#MergedPDF저장경로
def savemerged_path():
    savemerged_selected = filedialog.askdirectory()
    if savemerged_selected == '': #사용자가 취소를 누를 때
        return
    txt_mergedpdf_path.delete(0, END)
    txt_mergedpdf_path.insert(0,savemerged_selected)
    return savemerged_selected

# print WebtoPDF
def PrintSetUp():
    chrome_options = Options()
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    app_state = {
        "recentDestinations": [
            {
                "id": "Save as PDF",
                "origin": "local",
                "account": ""
            }
        ],
        "selectedDestinationId": "Save as PDF",
        "version": 2,
        "isLandscapeEnabled": False,
        "pageSize": 'A4',
        "scalingType": 3,
        "scaling": "40",
        "isHeaderFooterEnabled": False,
        "isCssBackgroundEnabled": True,
        "isColorEnabled": False,
        "isCollateEnabled": True
    }
    prefs = {
        'printing.print_preview_sticky_settings.appState': json.dumps(app_state),
        "download.default_directory": "C:/Users/swwoo/Downloads" # 원하는 다운로드 경로로 변경
    }
    chrome_options.add_experimental_option('prefs', prefs)
    chrome_options.add_argument('--kiosk-printing')
    return chrome_options


def WebToPDF(url):
    chrome_options = PrintSetUp()
    driver = webdriver.Chrome(options=chrome_options)
    driver.implicitly_wait(12)
    driver.get(url)
    wait = WebDriverWait(driver, 17)
    wait.until(EC.presence_of_all_elements_located)
    # time.sleep(2)
    driver.execute_script('return window.print()')
    time.sleep(5)
    driver.quit()


def start_process():
    adddrpath_selected = txt_frame1_path.get()
    xl_file_path = txt_xls_path.get()
    # xl_file_path = files    #'C:\\import\\HSCODE.xlsx'
    workbook = openpyxl.load_workbook(xl_file_path, data_only=True)
    worksheet = workbook.worksheets[0]

    for row, item in enumerate(worksheet.rows):
        if row > 0:
            url = item[1].value
            file_name = item[0].value

            try:
                WebToPDF(url)
        
                file_base = str(file_name) + ".pdf"
                file_count =+ 1
                filename = os.path.join(adddrpath_selected, f"{file_name} ({file_count}).pdf")
        
                while os.path.exists(filename):
                    file_count += 1
                    filename = os.path.join(adddrpath_selected, f"{file_name} ({file_count}).pdf")
            
                shutil.move(max([adddrpath_selected + "\\" + f for f in os.listdir(adddrpath_selected)], key=os.path.getctime), filename)  #'C:\\Users\\swwoo\\Downloads'

            except Exception as e:
                # print(f'Error processing URL {url}: {e}')
                print(url + '실패')

def run_thread():
    # 스레드 생성 및 시작
    thread = threading.Thread(target=start_process)
    thread.daemon = True
    thread.start()




def split_process():
# PDF 파일이 있는 폴더의 경로를 정의합니다.
    folder_path = txt_frame1_path.get()

# 폴더 내의 모든 파일을 반복합니다.
    for filename in os.listdir(folder_path):
        # PDF 파일만 처리합니다.
        if filename.endswith(".pdf"):
            # 파일의 전체 경로를 생성합니다.
            file_path = os.path.join(folder_path, filename)
            
            # PyPDF2를 사용하여 PDF 파일을 읽습니다.
            pdf_file = PyPDF2.PdfReader(file_path)
            
            # 수정된 버전을 저장할 새 PDF 파일을 생성합니다.
            output_pdf = PyPDF2.PdfWriter()
            
            # 새 PDF 파일에 1페이지를 추가합니다.
            output_pdf.add_page(pdf_file.pages[0])
        
            # 새 PDF 파일을 디스크에 쓰기
            with open(file_path, "wb") as output_file:
                output_pdf.write(output_file)


def check_pdf_excel():
    # PDF 파일이 있는 폴더 경로
    pdf_folder_path = txt_frame1_path.get()

    # 엑셀 파일 경로
    excel_file_path = txt_xls_path.get()

    # PDF 파일 이름과 경로를 딕셔너리로 저장
    pdf_dict = {}
    for pdf in os.listdir(pdf_folder_path):
        if pdf.endswith(".pdf"):
            pdf_name = os.path.splitext(pdf)[0] 
            pdf_path = os.path.join(pdf_folder_path, pdf)
            pdf_dict[pdf_name] = pdf_path

    # 엑셀 파일 열기
    workbook = load_workbook(filename=excel_file_path)
    worksheet = workbook.active

    # A 열에서 PDF 파일 이름 가져오기
    pdf_files = []
    for cell in worksheet['A'][1:]:
        count = 0
        filename = f"{os.path.splitext(cell.value)[0]} ({count+1}){os.path.splitext(cell.value)[1]}"
        while filename in pdf_files:
            count += 1
            filename = f"{os.path.splitext(cell.value)[0]} ({count+1}){os.path.splitext(cell.value)[1]}"
        pdf_files.append(filename)

    # C 열에서 PDF 파일 이름에 해당하는 값 가져오기
    for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            pdf_name = cell.offset(column=-2).value
            if pdf_name in pdf_files:
                pdf_path = pdf_dict.get(pdf_name)
                if pdf_path:
                    cell.value = pdf_files[pdf_files.index(pdf_name)]
                else:
                    cell.value = "" 
            else:
                cell.value = ""  

    workbook.save(excel_file_path)  




def merge_process():
    # 병합할 폴더 경로와 결과물 파일이 저장될 폴더 경로 설정
    folder_path = txt_frame1_path.get()
    result_folder_path = txt_mergedpdf_path.get()

    if not os.path.exists(result_folder_path):
        os.makedirs(result_folder_path)

    # 파일 이름의 앞 11자리가 동일한 파일들을 그룹화
    file_dict = defaultdict(list)
    for filename in os.listdir(folder_path):
        if filename.endswith('.pdf'):
            key = filename[:11]  # 앞 11자리까지만 추출하여 key로 사용
            file_dict[key].append(filename)

    # 그룹화된 파일들을 순회하며 각 그룹을 병합 또는 복사
    for key, filenames in file_dict.items():
        if len(filenames) == 1:
            file_path = os.path.join(folder_path, filenames[0])
            result_name = re.sub(r'\(\d+\)', '', filenames[0])  # 괄호와 괄호 안에 있는 숫자 삭제
            result_path = os.path.join(result_folder_path, result_name)
            copyfile(file_path, result_path)
        else:
            merger = PdfMerger()
            for filename in sorted(filenames):
                file_path = os.path.join(folder_path, filename)
                merger.append(file_path)
            result_name = filenames[0]
            # 파일 이름에서 괄호와 괄호 안에 있는 숫자를 삭제하여 결과물 파일 이름 생성
            result_name = re.sub(r'\(\d+\)', '', result_name)
            i = 1
            while os.path.exists(os.path.join(result_folder_path, result_name)):
                result_name = re.sub(r'\(\d+\)', '', result_name) + f'({i})'
                i += 1
            result_path = os.path.join(result_folder_path, result_name)
            merger.write(result_path)
            merger.close()



#프레임1 (기본 다운로드폴더 지정)
frame1 = LabelFrame(root, text="Select Download folder")
frame1.pack(fill="x", padx=5, pady=5, ipady=5)

txt_frame1_path = Entry(frame1)
txt_frame1_path.pack(side="left", fill="x", expand="True", ipady=4) #

btn_frame1_path = Button(frame1, text="...folder",padx=5, pady=5, width=12, command = add_path)
btn_frame1_path.pack(side="right")



#Web to PDF 저장
savepdf_frame = LabelFrame(root, text="Web to PDF")
savepdf_frame.pack(fill="x", padx=5, pady=5, ipady=5)

txt_xls_path = Entry(savepdf_frame)
txt_xls_path.pack(side="left", fill="x", expand="True", ipady=4) #

btn_start = Button(savepdf_frame, text="START",padx=5, pady=5, width=12, command=run_thread)
btn_start.pack(side="right")

btn_savepdf_path = Button(savepdf_frame, text="...xlsx file",padx=5, pady=5, width=12, command=add_file)
btn_savepdf_path.pack(side="right")



#Slit PDF
splitpdf_frame = LabelFrame(root, text="Split PDF")
splitpdf_frame.pack(fill="x",padx=5, pady=5, ipady=5)

btn_splitpdf = Button(splitpdf_frame, text="START", padx=5, pady=5, width="20", command=split_process)
btn_splitpdf.pack()



#Excel Check
xlscheck_frame = LabelFrame(root, text="Check PDF EXCEL")
xlscheck_frame.pack(fill="x",padx=5, pady=5, ipady=5)

btn_xlscheck = Button(xlscheck_frame, text="START", padx=5, pady=5, width="20", command=check_pdf_excel)
btn_xlscheck.pack()



# Merged PDF 저장경로
mergedpdf_frame = LabelFrame(root, text="Folder to merfed PDF")
mergedpdf_frame.pack(fill="x", padx=5, pady=5, ipady=5)

txt_mergedpdf_path = Entry(mergedpdf_frame)
txt_mergedpdf_path.pack(side="left", fill="x", expand="True", ipady=4) #

btn_mergestart = Button(mergedpdf_frame, text="START",padx=5, pady=5, width=12, command=merge_process)
btn_mergestart.pack(side="right")

btn_mergedpdf_path = Button(mergedpdf_frame, text="...Folder to save",padx=5, pady=5, width=12, command=savemerged_path)
btn_mergedpdf_path.pack(side="right")





# 진행상황 profress bar
frame_progress = LabelFrame(root, text="진행상황")
frame_progress.pack(fill="x")



# 종료 프레임
close_frame = Frame(root)
close_frame.pack(side="right", padx=5, pady=5)

btn_close = Button(close_frame, padx=5, pady=5, text="CLOSE", width=12)
btn_close.pack()


p_var = DoubleVar()
progress_bar = ttk.Progressbar(frame_progress, maximum=100, variable=p_var)
progress_bar.pack(fill="x")





root.mainloop()

