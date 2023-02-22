import os
from openpyxl import load_workbook

# PDF 파일이 있는 폴더 경로
pdf_folder_path = 'C:\\Users\\swwoo\\Downloads'

# 엑셀 파일 경로
excel_file_path = 'C:\\import\\HSCODE.xlsx'

# PDF 파일 이름과 경로를 딕셔너리로 저장
pdf_dict = {}
for pdf in os.listdir(pdf_folder_path):
    if pdf.endswith(".pdf"):
        pdf_name = os.path.splitext(pdf)[0].split("(")[0].strip()  # 파일 이름에서 (1), (2) 등의 부분 제외
        pdf_path = os.path.join(pdf_folder_path, pdf)
        pdf_dict[pdf_name] = pdf_path

# 엑셀 파일 열기
workbook = load_workbook(filename=excel_file_path)
worksheet = workbook.active

# A 열에서 PDF 파일 이름 가져오기
pdf_files = [cell.value for cell in worksheet['A'][1:]]

# C 열에서 PDF 파일 이름에 해당하는 값 가져오기
for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        pdf_name = cell.offset(column=-2).value.split("(")[0].strip()  # 파일 이름에서 (1), (2) 등의 부분 제외
        if pdf_name in pdf_files:
            pdf_path = pdf_dict.get(pdf_name)
            if pdf_path:
                cell.value = pdf_name  # A열과 동일한 PDF 파일이 있으면 A열의 값을 그대로 사용
            else:
                cell.value = ""  # A열과 동일한 PDF 파일이 없으면 C열을 공란으로 처리
        else:
            cell.value = ""  # A열과 동일한 PDF 파일이 없으면 C열을 공란으로 처리

workbook.save(excel_file_path)  # 변경된 내용을 저장