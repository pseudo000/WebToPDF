from selenium import webdriver
import time
options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_argument('--disable-gpu')
import openpyxl 

xl_file_path = 'C:\\import\\HSCODE.xlsx' 

wb =openpyxl.load_workbook(xl_file_path, data_only=True)  
ws = wb.worksheets[0]  

for row, item in enumerate(ws.rows):
    if(row > 0):   

        url_path = item[1].value  
        file_name = item[0].value 
        save_file_path = 'C:\\import\\' 


        for row in range(1, ws.max_row+1):
            file_name = ws.cell(row=row, column=1).value + ".pdf" 
            url_path = ws.cell(row=row, column=2).value  
            save_file_path = 'C:\\import\\' + file_name
            driver = webdriver.Chrome(options=options)
            driver.get(url_path)
            driver.execute_script('window.print();')
            time.sleep(3)
            driver.quit()